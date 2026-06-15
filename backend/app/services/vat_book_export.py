"""Stage D — write the VAT book to .xlsx in the format O'Farrell & Co receive.

Mirrors the operator's hand-made layout: a header block, one row per
transaction (Posted Account, Pub, Date, Descriptions, Debit/Credit, Type,
Category) with the gross amount placed in its VAT-band column (K-O), and a
totals row summing each band.
"""

from __future__ import annotations

import io
from decimal import Decimal

from app.models import Transaction
from app.services.vat_categorisation import VAT_BANDS, VatBookRow

# Column layout (matches the hand-made book).
_HEADERS = {
    "B": "Posted Account",
    "C": "Pub",
    "D": "Date",
    "E": "Description1",
    "F": "Description2",
    "G": "Debit Amount",
    "H": "Credit Amount",
    "I": "Transaction Type",
    "J": "Category",
    "K": "Resale @ 23%",
    "L": "Non-Resale @ 23%",
    "M": "Non-Resale @ 13.5%",
    "N": "Non-Resale @ 9%",
    "O": "Non-Resale @ 0%",
    "P": "Supporting Documents",
}
_BAND_COLUMN = {
    "resale_23": "K",
    "non_resale_23": "L",
    "non_resale_13_5": "M",
    "non_resale_9": "N",
    "non_resale_0": "O",
}


def _num(value: Decimal | None) -> float | None:
    return float(value) if value is not None else None


def _document_summary(documents: list[dict]) -> str:
    parts = []
    for doc in documents:
        bits = [doc.get("type", "").replace("_", " ").title()] if doc.get("type") else []
        if doc.get("supplier"):
            bits.append(doc["supplier"])
        if doc.get("reference"):
            bits.append(str(doc["reference"]))
        parts.append(" ".join(bits).strip())
    return "; ".join(p for p in parts if p)


def write_vat_book_xlsx(
    *,
    targets: list[Transaction],
    rows_by_id: dict,
    documents_by_transaction: dict | None = None,
    period_label: str,
) -> bytes:
    documents_by_transaction = documents_by_transaction or {}
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = f"VAT BOOK {period_label}"[:31]

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8EEF7")
    total_fill = PatternFill("solid", fgColor="F3F3F3")
    thin = Side(style="thin", color="D0D5DD")
    border = Border(bottom=thin)
    money_fmt = "#,##0.00"

    # Header row (row 1).
    for col, label in _HEADERS.items():
        cell = ws[f"{col}1"]
        cell.value = label
        cell.font = bold
        cell.fill = header_fill
        cell.border = border

    band_totals: dict[str, Decimal] = {key: Decimal("0.00") for key, _, _ in VAT_BANDS}
    r = 2
    for txn in targets:
        row: VatBookRow | None = rows_by_id.get(txn.id)
        ws[f"B{r}"] = txn.posted_account
        ws[f"C{r}"] = txn.pub
        ws[f"D{r}"] = txn.transaction_date
        if txn.transaction_date is not None:
            ws[f"D{r}"].number_format = "dd/mm/yyyy"
        ws[f"E{r}"] = txn.description1
        ws[f"F{r}"] = txn.description2
        ws[f"G{r}"] = _num(txn.debit_amount)
        ws[f"G{r}"].number_format = money_fmt
        ws[f"H{r}"] = _num(txn.credit_amount)
        ws[f"H{r}"].number_format = money_fmt
        ws[f"I{r}"] = txn.transaction_type
        ws[f"J{r}"] = row.predicted_category if row else txn.category

        band = row.predicted_band if row else None
        if band and band in _BAND_COLUMN and txn.debit_amount is not None:
            column = _BAND_COLUMN[band]
            ws[f"{column}{r}"] = _num(txn.debit_amount)
            ws[f"{column}{r}"].number_format = money_fmt
            band_totals[band] += txn.debit_amount
        ws[f"P{r}"] = _document_summary(documents_by_transaction.get(txn.id, []))
        r += 1

    # Totals row.
    ws[f"J{r}"] = "TOTALS"
    ws[f"J{r}"].font = bold
    for key, _, _ in VAT_BANDS:
        column = _BAND_COLUMN[key]
        cell = ws[f"{column}{r}"]
        cell.value = float(band_totals[key])
        cell.number_format = money_fmt
        cell.font = bold
        cell.fill = total_fill

    widths = {"B": 18, "C": 9, "D": 11, "E": 24, "F": 22, "G": 13, "H": 13, "I": 14, "J": 26}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for col in _BAND_COLUMN.values():
        ws.column_dimensions[col].width = 15
    ws.column_dimensions["P"].width = 40
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
